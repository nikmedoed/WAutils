from selenium import webdriver
from time import sleep
from openpyxl import Workbook

def finder(base):
    result = []
    browser = webdriver.Chrome()
    browser.get('https://web.whatsapp.com/')
    input("enter как авторизируешься")
    search = browser.find_element_by_xpath('//div[1]/div/label/input')
    for i in base:
        search.clear()
        try:
            search.send_keys(i[2])
            sleep(2)
            pers = browser.find_elements_by_xpath('//div[2]/div[1]/div/span/span/span')
            pers[0].click()
            sleep(6) # готовы читать
            mess = browser.find_elements_by_xpath('//div[3]/div/div/div[2]/div')
            if len(mess) > 0:
                mes = "\n------\n".join(list(map(lambda x: x.text, mess[2:])))
                date = browser.find_elements_by_xpath('//div[3]/div/div/div[2]/div/div/div/div/div[1]')
                dat = date[len(date)-1].get_attribute('data-pre-plain-text').split()[1].replace(']', "")
                print(i, dat)
                i.extend([mes, dat])
                result.append(i)
            else:
                print("Пусто", i)
                i.append("Пусто")
                result.append(i)
        except:
            print("Исключение", i)
            i.append("Error")
            result.append(i)
    browser.close()
    return result


if __name__ == "__main__":
    with open("../OKbase1.txt", "r", encoding="utf8") as b:
        base = list(map(lambda x: x.split("\t"), b.read().split("\n")))
    fin = finder(base)
    print(fin)
    wb = Workbook()
    ws = wb.active
    ws.title = "WA messages"
    for i in range(len(fin)):
        for j in range(len(fin[i])):
            # print(fin[i][j])
            ws.cell(row=i + 1, column=j + 1).value = fin[i][j]
        if "Это так?" in fin[i][j]:
            ws.cell(row=i + 1, column=j + 2).value = 'NoAnsw'
    wb.save('..\\findedMessages.xlsx')

